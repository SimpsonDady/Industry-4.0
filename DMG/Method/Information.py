from openpyxl import Workbook, load_workbook


class timemode:
    def __init__(self, machine_name, inputdata):
        self.machine_name = machine_name
        self.inputdata = inputdata
        self.ws = Workbook()
        self.wb = self.ws.active
        self.open = load_workbook('DMG01_0703-0707.xlsx')
        self.xlsx = self.open['Sheet1']
        self.output()

    def output(self):
        self.wb.append(['工具代碼', '件號', '工號', 'NC程式', '版別', '工作中心', '機台編號', '開始日期', '開始時間', '結束日期', '結束時間', '加工工時'])
        code = self.inputdata[0].getProgramCode(0)
        date = self.inputdata[0].getDate()
        starttime = self.inputdata[0].getStartTime(0)
        startday = self.inputdata[0].getDay()
        startsecond = self.inputdata[0].getStartSecond(0)
        endtime = self.inputdata[0].getEndTime(0)
        endday = self.inputdata[0].getDay()
        endsecond = self.inputdata[0].getEndSecond(0)
        worknum = 'N/A'
        NC = 'N/A'
        workcenter = 'N/A'
        component = 'N/A'
        for i in range(2, self.xlsx.max_row):
            if self.xlsx['G' + str(i)].value == self.inputdata[0].getComponent(0) and self.xlsx['M' + str(i)].value == self.inputdata[0].getProgramCode(0):
                print(21354)
                worknum = self.xlsx['B' + str(i)].value
                NC = self.xlsx['J' + str(i)].value
                workcenter = self.xlsx['D' + str(i)].value
                component = self.xlsx['G' + str(i)].value
        temp = [self.inputdata[0].getProgramCode(0), component, worknum, NC,
                self.inputdata[0].getVersion(0), workcenter, self.machine_name]
        for i in self.inputdata:
            for j in range(i.getProgramLength()):
                if i.getProgramCode(j) == 'TOOL-SL-D.I' or i.getProgramCode(j) == 'WARM' or i.getProgramCode(j) == code:
                    starttime = i.getStartTime(j)
                    startday = i.getDay()
                    startsecond = i.getStartSecond(j)
                    if date != i.getDate():
                        endsecond += 86400
                    date = i.getDate()
                else:
                    worktime = (endsecond - startsecond) / 3600
                    temp.extend([startday, starttime, endday, endtime, worktime])
                    self.wb.append(temp)
                    temp = []
                    startday = i.getDay()
                    starttime = i.getStartTime(j)
                    startsecond = i.getStartSecond(j)
                    code = i.getProgramCode(j)
                    endtime = i.getEndTime(j)
                    endday = i.getDay()
                    endsecond = i.getEndSecond(j)
                    worknum = 'N/A'
                    NC = 'N/A'
                    workcenter = 'N/A'
                    component = 'N/A'
                    for k in range(2, self.xlsx.max_row):
                        if self.xlsx['G' + str(k)].value == i.getComponent(j) and self.xlsx[
                                    'M' + str(k)].value == i.getProgramCode(j):
                            print(21354)
                            worknum = self.xlsx['B' + str(k)].value
                            NC = self.xlsx['J' + str(k)].value
                            workcenter = self.xlsx['D' + str(k)].value
                            component = self.xlsx['G' + str(k)].value
                    temp = [i.getProgramCode(j), component, worknum, NC, i.getVersion(j), workcenter, self.machine_name]
        worktime = (endsecond - startsecond) / 3600
        temp.extend([startday, starttime, endday, endtime, worktime])
        self.wb.append(temp)
        self.ws.save(self.machine_name + '_TimeModel.xlsx')
