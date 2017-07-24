from openpyxl import Workbook, load_workbook


class Timesmode:
    def __init__(self, plan_file, machine_name, inputdata):
        self.machine_name = machine_name
        self.inputdata = inputdata
        self.ws = Workbook()
        self.wb = self.ws.active
        self.open = load_workbook(plan_file)
        self.xlsx = self.open['Sheet1']
        self.wb.append(['工具代碼', '件號', '工號', 'NC程式', '版別', '工作中心', '機台編號',
                        '開始日期', '開始時間', '結束日期', '結束時間', '加工工時'])
        self.output()

    def output(self):
        warm = 0
        startcode = 0
        startdate = 0
        endcode = 0
        enddate = 0
        totaltime = 0
        for i in range(len(self.inputdata)):
            for j in range(self.inputdata[i].getProgramLength()):
                if self.inputdata[i].getProgramCode(j) == 'TOOL-SL-D':
                    continue
                elif self.inputdata[i].getProgramCode(j) == 'WARM':
                    warm = 2
                    self.save(startdate, startcode, enddate, endcode)
                elif self.inputdata[i].getProgramCode(j) == self.inputdata[enddate].getProgramCode(endcode) and not warm:
                    startdate = i
                    startcode = j
                elif warm:
                    startdate = i
                    startcode = j
                    enddate = i
                    endcode = j
                    warm -= 1
                    if warm:
                        self.save(startdate, startcode, enddate, endcode)
                else:
                    self.save(startdate, startcode, enddate, endcode)
                    startdate = i
                    startcode = j
                    enddate = i
                    endcode = j
        self.save(startdate, startcode, enddate, endcode)
        self.ws.save('C:\\Users\\User\\Desktop\\result\\time_model\\' + self.machine_name + '_TimeModel.xlsx')

    def save(self, startdate, startcode, enddate, endcode):
        worknum = 'N/A'
        NC = 'N/A'
        workcenter = 'N/A'
        component = 'N/A'
        for i in range(2, self.xlsx.max_row):
            if self.xlsx['G' + str(i)].value == self.inputdata[startdate].getComponent(startcode) \
                    and self.xlsx['M' + str(i)].value == self.inputdata[startdate].getProgramCode(startcode):
                worknum = self.xlsx['B' + str(i)].value
                NC = self.xlsx['J' + str(i)].value
                workcenter = self.xlsx['D' + str(i)].value
                component = self.xlsx['G' + str(i)].value
        time = (self.inputdata[enddate].getEndSecond(endcode)-self.inputdata[startdate].getStartSecond(startcode))/3600
        if time < 0:
            time += 86400
        self.wb.append([self.inputdata[startdate].getProgramCode(startcode), component, worknum, NC,
                        self.inputdata[startdate].getVersion(startcode), workcenter, self.machine_name,
                        self.inputdata[startdate].getDay(), self.inputdata[startdate].getStartTime(startcode),
                        self.inputdata[enddate].getDay(), self.inputdata[enddate].getEndTime(endcode),
                        time])
