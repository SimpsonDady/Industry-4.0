from openpyxl import load_workbook  # for read Excel


class Cutworklist:
    def __init__(self, filename):
        self.start_date = []    # list of {'year', 'month', 'date'}
        self.start_time = []    # list of {'hour', 'minute', 'second'}
        self.end_date = []      # list of {'year', 'month', 'date'}
        self.end_time = []      # list of {'hour', 'minute', 'second'}
        self.code = []          # list of main program code
        self.split(filename)

    def split(self, filename):
        xlsx = load_workbook(filename)          # load Excel
        xlsx = xlsx['Sheet1']                   # Excel page
        for line in range(2, xlsx.max_row + 1):             # read row (line)
            for section in range(1, xlsx.max_column + 1):   # read column
                data = str(xlsx.cell(row=line, column=section).value)   # read one data
                if section == 5 or section == 6:        # Start time(5), End time(6)
                    data_cut = data.split(' ')          # Split date, time
                    date_cut = data_cut[0].split('-')   # Split year, month, date
                    time_cut = data_cut[1].split(':')   # Split hour, minute, second
                    if section == 5:                    # Save start time(5)
                        self.start_date.append({'year': int(date_cut[0]), 'month': int(date_cut[1]), 'date': int(date_cut[2])})
                        self.start_time.append({'hour': int(time_cut[0]), 'minute': int(time_cut[1]), 'second': int(time_cut[2])})
                    if section == 6:                    # Save end time(6)
                        self.end_date.append({'year': int(date_cut[0]), 'month': int(date_cut[1]), 'date': int(date_cut[2])})
                        self.end_time.append({'hour': int(time_cut[0]), 'minute': int(time_cut[1]), 'second': int(time_cut[2])})
                if section == 13:                       # main program code(13)
                    self.code.append(data)
