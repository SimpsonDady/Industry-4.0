from datetime import datetime

from openpyxl import load_workbook

from method.DataStructure import Data, Word


class ReadFile:
    def __init__(self):
        self.component = []
        self.kinfe = []
        self.startTime = []
        self.endTime = []
        self.Data = []
        self.readfile()
    def readfile(self):
        file = load_workbook(filename=r'D:\\result\\time_model\\DMG01_TimeModel.xlsx')
        xlsx = file['Sheet']
        for line in range(2, xlsx.max_row + 1):  # read row (line)
            for section in range(1, xlsx.max_column + 1):  # read column
                data = str(xlsx.cell(row=line, column=section).value)
                if section == 1:
                    self.component.append(data)
                if section == 2:
                    self.kinfe.append(data)
                if section == 3:
                    try:
                        self.startTime.append(datetime.strptime(data, "%Y-%m-%d %H:%M:%S"))
                    except ValueError:
                        data = data[:25]
                        self.startTime.append(datetime.strptime(data, "%Y-%m-%d %H:%M:%S.%f"))
                if section == 4:
                    try:
                        self.endTime.append(datetime.strptime(data, "%Y-%m-%d %H:%M:%S"))
                    except ValueError:
                        data = data[:25]
                        self.endTime.append(datetime.strptime(data, "%Y-%m-%d %H:%M:%S.%f"))
        pre_component = self.component[0]
        pre_knife = self.kinfe[0]
        word = Word()
        data = Data(self.component[0])
        componentList = []
        self.Data.append(data)
        componentList.append(self.component[0])
        for i in range(len(self.component)):
            if self.component[i] != pre_component:
                for j in self.Data:
                    if j.component == pre_component:
                        j.add_text(word.word)
                        if self.kinfe[i] == pre_knife:
                            word.word[-1].filterlast = True
                        word = Word()
                        pre_component = self.component[i]
                        break
                if self.component[i] not in componentList:
                    data = Data(self.component[i])
                    self.Data.append(data)
                    pre_component = self.component[i]
                    componentList.append(self.component[i])
                    word = Word()
            pre_knife = self.kinfe[i]
            word.add_knife(self.kinfe[i], self.startTime[i], self.endTime[i], False)
            if i == len(self.component)-1:
                for j in self.Data:
                    if j.component == pre_component:
                        j.add_text(word.word)
                        break